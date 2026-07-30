"""
Migrate REVISED_DETAILED_SCHEDULE.xlsx into the Retreat Ops tables
(run_of_show_blocks, roster_teams, setup_tasks, contingency_plans).

Reads four sheets from the workbook:
  - "Internal Schedule"   -> RunOfShowBlock  (one section per day: FRIDAY / SATURDAY / SUNDAY)
  - "Team Roster"         -> RosterTeam
  - "Setup & Supplies"    -> SetupTask
  - "Contingency"         -> ContingencyPlan

Usage (run from the backend/ directory, same folder as app.py):
    python3 migrate_retreat_ops.py /path/to/REVISED_DETAILED_SCHEDULE.xlsx
    python3 migrate_retreat_ops.py /path/to/REVISED_DETAILED_SCHEDULE.xlsx --clear
    python3 migrate_retreat_ops.py /path/to/SCHEDULE.xlsx --days "Thursday,Friday,Saturday,Sunday"

--clear wipes any existing Retreat Ops rows before importing (use on a fresh
import / re-import; without it, running the script twice will duplicate rows).

--days lets you tell the script which day names to expect in the "Internal
Schedule" sheet's section headers (default: Friday,Saturday,Sunday). If the
sheet contains a day section that doesn't match this list, the script stops
and tells you exactly which header it didn't recognize, instead of silently
dropping that day's blocks — add the missing day to --days and re-run.

This uses the Flask app + SQLAlchemy models directly (not the HTTP API), so it
works against whatever database your .env / config.py already points to
(SQLite locally, MySQL in production) without needing a running server or a
login token.
"""

import sys
import os
import argparse
import datetime

import openpyxl

# --- Load the Flask app & models -------------------------------------------------
sys.path.insert(0, ".")
from app import create_app
from db import db
from models import RunOfShowBlock, RosterTeam, SetupTask, ContingencyPlan


def fmt_time(value):
    """Convert an openpyxl cell value (datetime.time / datetime.datetime / str / None) to 'hh:mm AM/PM'."""
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.strftime("%I:%M %p")
    if isinstance(value, datetime.time):
        return value.strftime("%I:%M %p")
    
    text = str(value).strip()
    if not text:
        return None
        
    for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M%p"):
        try:
            dt = datetime.datetime.strptime(text, fmt)
            return dt.strftime("%I:%M %p")
        except ValueError:
            continue
            
    return text


def clean(value):
    """Strip strings, turn blank/'—' placeholders into None."""
    if value is None:
        return None
    text = str(value).strip()
    if text == "" or text == "—":
        return None
    return text


def parse_internal_schedule(ws, valid_days):
    """
    Walk the "Internal Schedule" sheet. Structure is:
      Row 1: workbook-wide title
      Then repeating per day:
        - a day-divider row, e.g. "FRIDAY  —  Arrival, Session 1, Campfire"
        - a column-header row starting with "Start"
        - N data rows
        - a blank row

    valid_days: list of day names to recognize (case-insensitive match against
    the first word of each divider row). Any divider whose day name doesn't
    match one of these is reported back in `unmatched_headers` rather than
    silently skipped, so a renamed/added day surfaces instead of vanishing.
    """
    valid_days_upper = {d.upper(): d for d in valid_days}
    blocks = []
    unmatched_headers = []
    current_day = None
    order_index = 0

    for row in ws.iter_rows(min_row=2, values_only=True):  # row 1 is always the workbook-wide title
        first_cell = row[0]

        if first_cell is None:
            continue  # blank separator row

        if isinstance(first_cell, str) and first_cell.strip().upper() == "START":
            continue  # column header row

        if isinstance(first_cell, str) and (row[1] is None):
            # A day-divider row: first cell holds free text, rest of the row is empty.
            day_text = first_cell.strip()
            if "—" in day_text:
                day_text = day_text.split("—")[0]
            elif "-" in day_text:
                day_text = day_text.split("-")[0]
            day_token = day_text.split()[0] if day_text.split() else day_text

            if day_token.upper() in valid_days_upper:
                current_day = valid_days_upper[day_token.upper()]
                order_index = 0
            else:
                unmatched_headers.append(first_cell.strip())
                current_day = None
            continue

        if current_day is None:
            continue  # still inside the title / preamble, or an unmatched day section

        # A real data row: Start, End, Duration, Block/Event, Location, Point Person,
        # Supporting Teams, Setup Done By, Setup/Teardown Notes, Tech & Production Cues,
        # Kidz Corner, Contingency, Status
        (start, end, _duration, block_title, location, point_person, supporting_teams,
         setup_time, setup_notes, tech_cues, kidz_corner, contingency, status) = (
            list(row) + [None] * (13 - len(row))
        )[:13]

        if not clean(block_title):
            continue

        blocks.append(RunOfShowBlock(
            day=current_day,
            order_index=order_index,
            start_time=fmt_time(start),
            end_time=fmt_time(end),
            block_title=clean(block_title),
            location=clean(location),
            point_person=clean(point_person),
            supporting_teams=clean(supporting_teams),
            setup_time=fmt_time(setup_time),
            setup_notes=clean(setup_notes),
            tech_cues=clean(tech_cues),
            kidz_corner_note=clean(kidz_corner),
            contingency=clean(contingency),
            status=clean(status) or "Not started",
        ))
        order_index += 1

    return blocks, unmatched_headers


def parse_team_roster(ws):
    teams = []
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[2:]:  # skip title row + header row
        name = clean(row[0]) if len(row) > 0 else None
        if not name:
            continue
        teams.append(RosterTeam(
            name=name,
            lead=clean(row[1]) if len(row) > 1 else None,
            phone=clean(row[2]) if len(row) > 2 else None,
            members=clean(row[3]) if len(row) > 3 else None,
            owns_blocks=clean(row[4]) if len(row) > 4 else None,
            notes=clean(row[5]) if len(row) > 5 else None,
        ))
    return teams


def parse_setup_supplies(ws):
    tasks = []
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[2:]:  # skip title row + header row
        item = clean(row[0]) if len(row) > 0 else None
        if not item:
            continue
        done_val = row[5] if len(row) > 5 else None
        done = str(done_val).strip().lower() in ("x", "yes", "true", "done", "1") if done_val else False
        tasks.append(SetupTask(
            item=item,
            for_block=clean(row[1]) if len(row) > 1 else None,
            owner=clean(row[2]) if len(row) > 2 else None,
            deadline=clean(row[3]) if len(row) > 3 else None,
            qty_detail=clean(row[4]) if len(row) > 4 else None,
            done=done,
        ))
    return tasks


def parse_contingency(ws):
    plans = []
    rows = list(ws.iter_rows(values_only=True))
    order_index = 0
    for row in rows[2:]:  # skip title row + header row
        scenario = clean(row[0]) if len(row) > 0 else None
        if not scenario:
            continue
        plans.append(ContingencyPlan(
            scenario=scenario,
            trigger=clean(row[1]) if len(row) > 1 else None,
            action=clean(row[2]) if len(row) > 2 else None,
            who_decides=clean(row[3]) if len(row) > 3 else None,
            order_index=order_index,
        ))
        order_index += 1
    return plans


def migrate(xlsx_path, clear=False, days=None):
    days = days or ["Friday", "Saturday", "Sunday"]

    # 1. Load the workbook FIRST so a bad file never touches the database.
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        print(f"Error: could not open '{xlsx_path}': {e}")
        return

    required_sheets = ["Internal Schedule", "Team Roster", "Setup & Supplies", "Contingency"]
    missing = [s for s in required_sheets if s not in wb.sheetnames]
    if missing:
        print(f"Error: workbook is missing expected sheet(s): {missing}. Aborting.")
        return

    blocks, unmatched_headers = parse_internal_schedule(wb["Internal Schedule"], days)

    if unmatched_headers:
        print("Error: found day section(s) in 'Internal Schedule' that don't match any "
              f"known day ({', '.join(days)}):")
        for h in unmatched_headers:
            print(f"    - {h!r}")
        print("Add the missing day with --days, e.g.:")
        print(f'    --days "{",".join(days)},Monday"')
        print("Aborting without touching the database.")
        return

    teams = parse_team_roster(wb["Team Roster"])
    tasks = parse_setup_supplies(wb["Setup & Supplies"])
    plans = parse_contingency(wb["Contingency"])

    print(f"Parsed from workbook: {len(blocks)} schedule blocks, {len(teams)} teams, "
          f"{len(tasks)} setup tasks, {len(plans)} contingency scenarios.")

    if not any([blocks, teams, tasks, plans]):
        print("Nothing parsed from the workbook — aborting without touching the database.")
        return

    # 2. Only now touch the database.
    app = create_app()
    with app.app_context():
        if clear:
            deleted = (
                RunOfShowBlock.query.delete(),
                RosterTeam.query.delete(),
                SetupTask.query.delete(),
                ContingencyPlan.query.delete(),
            )
            db.session.commit()
            print(f"Cleared existing rows: {deleted[0]} blocks, {deleted[1]} teams, "
                  f"{deleted[2]} tasks, {deleted[3]} plans.")

        db.session.add_all(blocks)
        db.session.add_all(teams)
        db.session.add_all(tasks)
        db.session.add_all(plans)
        db.session.commit()

    print("Migration complete.")
    print(f"  Run of Show blocks: {len(blocks)}")
    print(f"  Roster teams:       {len(teams)}")
    print(f"  Setup tasks:        {len(tasks)}")
    print(f"  Contingency plans:  {len(plans)}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Migrate the retreat schedule Excel workbook into Retreat Ops.")
    parser.add_argument("xlsx_path", help="Path to REVISED_DETAILED_SCHEDULE.xlsx")
    parser.add_argument("--clear", action="store_true",
                         help="Delete existing Retreat Ops rows before importing (use for a clean re-import).")
    parser.add_argument("--days", default="Friday,Saturday,Sunday",
                         help="Comma-separated list of day names to recognize in the 'Internal Schedule' "
                              "sheet's section headers (default: Friday,Saturday,Sunday). Add to this list "
                              "if a future retreat has more/different days, e.g. "
                              '--days "Thursday,Friday,Saturday,Sunday"')
    args = parser.parse_args()
    day_list = [d.strip() for d in args.days.split(",") if d.strip()]
    migrate(args.xlsx_path, clear=args.clear, days=day_list)
