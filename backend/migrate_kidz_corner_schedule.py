import os
import openpyxl
from app import create_app
from db import db
from models import (
    KidzCornerScheduleItem,
    KidzCornerCraft,
    KidzCornerAVLink,
    KidzCornerVolunteer,
    KidzCornerKid,
    KidzCornerBudgetItem,
)

# 1. Excel File Path
excel_path = r"C:\Users\jenyj\Downloads\Kidz Corner Schedule.xlsx"

# 2. Check if file exists
if not os.path.exists(excel_path):
    # Fallback to local script relative path if moved
    excel_path = os.path.join(os.path.dirname(__file__), "Kidz Corner Schedule.xlsx")

print(f"Reading schedule data from: {excel_path}")

if not os.path.exists(excel_path):
    print("CRITICAL ERROR: Excel file 'Kidz Corner Schedule.xlsx' not found!")
    exit(1)

# 3. Load Excel Workbook
wb = openpyxl.load_workbook(excel_path, read_only=True)

# 4. Initialize Flask context
app = create_app()

with app.app_context():
    print("Connected to database. Clearing old Kidz Corner data...")
    
    # Clean previous data to prevent duplicates
    try:
        db.session.query(KidzCornerScheduleItem).delete()
        db.session.query(KidzCornerCraft).delete()
        db.session.query(KidzCornerAVLink).delete()
        db.session.query(KidzCornerVolunteer).delete()
        db.session.query(KidzCornerKid).delete()
        db.session.query(KidzCornerBudgetItem).delete()
        db.session.commit()
        print("Cleared previous schedule items, crafts, AV links, volunteers, kids, and budget items.")
    except Exception as e:
        db.session.rollback()
        print(f"Error clearing previous data: {e}")
        exit(1)

    # 5. Parse People Sheet (Volunteers & Kids rosters)
    if "People" in wb.sheetnames:
        print("Parsing sheet: People")
        sheet = wb["People"]
        rows = list(sheet.iter_rows(values_only=True))
        
        # Parse Volunteers (Columns A & B) starting at row 3 (index 2)
        v_order = 0
        for r_idx in range(2, len(rows)):
            row = rows[r_idx]
            if not row:
                continue
            name_val = row[0]
            assign_val = row[1] if len(row) > 1 else None
            
            if name_val and str(name_val).strip():
                name_str = str(name_val).strip()
                if name_str.lower().startswith("name") or "volunteers" in name_str.lower():
                    continue
                v = KidzCornerVolunteer(
                    name=name_str,
                    assignment=str(assign_val).strip() if assign_val else None,
                    order_index=v_order
                )
                db.session.add(v)
                v_order += 1
                
        # Parse Kids (Columns D, E & F) starting at row 3 (index 2)
        k_order = 0
        for r_idx in range(2, len(rows)):
            row = rows[r_idx]
            if not row or len(row) < 4:
                continue
            name_val = row[3]
            age_val = row[4] if len(row) > 4 else None
            allergy_val = row[5] if len(row) > 5 else None
            
            if name_val and str(name_val).strip():
                name_str = str(name_val).strip()
                if name_str.lower().startswith("name") or "kids" in name_str.lower():
                    continue
                
                # Parse Age
                age_num = None
                if age_val is not None:
                    try:
                        age_num = int(age_val)
                    except (ValueError, TypeError):
                        pass
                
                k = KidzCornerKid(
                    name=name_str,
                    age=age_num,
                    allergies=str(allergy_val).strip() if allergy_val else None,
                    order_index=k_order
                )
                db.session.add(k)
                k_order += 1
        print(f"Migrated {v_order} volunteers and {k_order} kids from People sheet.")

    # 6. Parse Budget.Expenses Sheet
    if "Budget.Expenses" in wb.sheetnames:
        print("Parsing sheet: Budget.Expenses")
        sheet = wb["Budget.Expenses"]
        rows = list(sheet.iter_rows(values_only=True))
        
        b_order = 0
        for r_idx in range(1, len(rows)):
            row = rows[r_idx]
            if not row:
                continue
                
            month_val = row[0]
            inc_val = row[1] if len(row) > 1 else None
            exp_val = row[2] if len(row) > 2 else None
            proj_val = row[3] if len(row) > 3 else None
            file_val = row[4] if len(row) > 4 else None
            notes_val = row[5] if len(row) > 5 else None
            
            # Skip header row if it occurs
            if month_val and str(month_val).strip().lower().startswith("month"):
                continue
                
            # Skip SUM formula rows
            if inc_val and str(inc_val).strip().startswith("="):
                continue
            if exp_val and str(exp_val).strip().startswith("="):
                continue
            if proj_val and str(proj_val).strip().startswith("="):
                continue
                
            # Save if at least one value is set
            if any([month_val, inc_val, exp_val, proj_val, notes_val]):
                def to_float(v):
                    if v is None:
                        return None
                    try:
                        return float(v)
                    except (ValueError, TypeError):
                        return None
                        
                b_item = KidzCornerBudgetItem(
                    month=str(month_val).strip() if month_val else None,
                    income_actual=to_float(inc_val),
                    expenses_actual=to_float(exp_val),
                    expenses_projected=to_float(proj_val),
                    related_files=str(file_val).strip() if file_val else None,
                    notes=str(notes_val).strip() if notes_val else None,
                    order_index=b_order
                )
                db.session.add(b_item)
                b_order += 1
        print(f"Migrated {b_order} budget expense line items from Budget.Expenses sheet.")

    # 7. Parse Schedule & Craft Sheets
    for sheet_name in ["Friday Night", "Saturday Morning", "Saturday Night", "Sunday Morning"]:
        if sheet_name not in wb.sheetnames:
            print(f"Skipping sheet '{sheet_name}' (not found in workbook).")
            continue
            
        print(f"Parsing sheet: {sheet_name}")
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        
        order_idx = 0
        craft_title = None
        craft_materials = None
        craft_howto = None
        craft_ages = None
        
        for row in rows:
            if not row or len(row) < 2:
                continue
                
            time_val = row[0]
            activity_val = row[1]
            vol_val = row[2] if len(row) > 2 else None
            items_val = row[3] if len(row) > 3 else None
            notes_val = row[4] if len(row) > 4 else None
            
            if time_val:
                time_str = str(time_val).strip()
                
                # Check for headers or titles to skip
                if time_str.lower().startswith("time"):
                    continue
                if "schedule" in time_str.lower():
                    continue
                
                # Parse Craft details in Column A
                if time_str.startswith("Craft:"):
                    craft_title = time_str.replace("Craft:", "").strip()
                    continue
                if time_str.startswith("Materials:"):
                    craft_materials = time_str.replace("Materials:", "").strip()
                    continue
                if time_str.startswith("How-to:"):
                    craft_howto = time_str.replace("How-to:", "").strip()
                    continue
                if time_str.startswith("Ages:"):
                    craft_ages = time_str.replace("Ages:", "").strip()
                    continue
                
                # Insert schedule items
                if activity_val:
                    item = KidzCornerScheduleItem(
                        day=sheet_name,
                        time=time_str,
                        activity=str(activity_val).strip(),
                        volunteers_needed=str(vol_val).strip() if vol_val else None,
                        items_needed=str(items_val).strip() if items_val else None,
                        notes=str(notes_val).strip() if notes_val else None,
                        order_index=order_idx
                    )
                    db.session.add(item)
                    order_idx += 1

        # Save parsed craft details for this session if present
        if craft_title or craft_materials or craft_howto or craft_ages:
            craft = KidzCornerCraft(
                day=sheet_name,
                title=craft_title or "Kidz Corner Craft",
                materials=craft_materials,
                how_to=craft_howto,
                ages=craft_ages,
                order_index=0
            )
            db.session.add(craft)
            print(f"Added Craft details for {sheet_name}: {craft.title}")

    # 8. Parse AudioVideo Sheet
    if "AudioVideo" in wb.sheetnames:
        print("Parsing sheet: AudioVideo")
        sheet = wb["AudioVideo"]
        rows = list(sheet.iter_rows(values_only=True))
        
        category = "Action Song List"
        order_idx = 0
        
        for row in rows:
            if not row or len(row) < 2:
                continue
                
            col1 = row[0]
            col2 = row[1]
            
            if col1:
                col1_str = str(col1).strip()
                if "song list" in col1_str.lower():
                    category = "Action Song List"
                    continue
                if "dance party" in col1_str.lower():
                    category = "Dance Party"
                    continue
                
                if col2:
                    url_str = str(col2).strip()
                    if url_str.startswith("http"):
                        label = col1_str.rstrip(" -:")
                        av_link = KidzCornerAVLink(
                            category=category,
                            label=label,
                            url=url_str,
                            order_index=order_idx
                        )
                        db.session.add(av_link)
                        order_idx += 1
                        print(f"Added AV Link: {label} ({category})")

    # 9. Commit changes
    try:
        db.session.commit()
        print("SUCCESS: Kidz Corner database migration completed successfully!")
    except Exception as e:
        db.session.rollback()
        print(f"CRITICAL ERROR committing data to database: {e}")
        exit(1)
