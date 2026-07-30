from flask import Blueprint, request, jsonify
from flask_jwt_extended import jwt_required, get_jwt
from models import Camper, Tshirt
from db import db
from utils.limiter import rate_limit

from utils.permissions import require_page_permission

campers_bp = Blueprint("campers", __name__)

def require_admin():
    claims = get_jwt()
    return claims.get("role") in ["admin", "owner"]

def check_permission(role, page_key, required_level="read"):
    from models.permission import PagePermission
    from routes.permissions import DEFAULT_PERMISSIONS
    if role == "owner":
        return True
    access_level = DEFAULT_PERMISSIONS.get(role, {}).get(page_key, "hide")
    custom_perm = PagePermission.query.filter_by(role=role, page_key=page_key).first()
    if custom_perm:
        access_level = custom_perm.access_level
    if required_level == "edit" and access_level != "edit":
        return False
    if required_level == "read" and access_level == "hide":
        return False
    return True

@campers_bp.route("/", methods=["GET"])
@jwt_required()
def get_campers():
    claims = get_jwt()
    role = claims.get("role", "user")
    if not check_permission(role, "campers", "read") and not check_permission(role, "teams", "read") and not check_permission(role, "apparel", "read"):
        return jsonify({"error": "Access denied"}), 403
    search = request.args.get("search", "").strip()
    status = request.args.get("status", "").strip()
    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 20))

    query = Camper.query

    if search:
        like = f"%{search}%"
        # Find all family groups for campers matching the search criteria
        matched_families = db.session.query(Camper.family_group).filter(
            db.or_(
                Camper.first_name.ilike(like),
                Camper.last_name.ilike(like),
                Camper.guardian_name.ilike(like),
                Camper.family_group.ilike(like),
            )
        ).filter(Camper.family_group.isnot(None), Camper.family_group != '').distinct().all()
        
        family_ids = [f[0] for f in matched_families if f[0]]
        
        if family_ids:
            query = query.filter(
                db.or_(
                    Camper.first_name.ilike(like),
                    Camper.last_name.ilike(like),
                    Camper.guardian_name.ilike(like),
                    Camper.family_group.ilike(like),
                    Camper.family_group.in_(family_ids)
                )
            )
        else:
            query = query.filter(
                db.or_(
                    Camper.first_name.ilike(like),
                    Camper.last_name.ilike(like),
                    Camper.guardian_name.ilike(like),
                    Camper.family_group.ilike(like),
                )
            )
    if status:
        query = query.filter(Camper.registration_status == status)

    order_by_clause = db.case(
        (db.or_(Camper.family_group.is_(None), Camper.family_group == ''), 1),
        else_=0
    )
    ordered_query = query.order_by(
        order_by_clause,
        Camper.family_group,
        Camper.last_name,
        Camper.first_name
    )

    if per_page == -1:
        items = ordered_query.all()
        return jsonify({
            "campers": [c.to_dict() for c in items],
            "total": len(items),
            "pages": 1,
            "page": 1,
        }), 200

    paginated = ordered_query.paginate(
        page=page, per_page=per_page, error_out=False
    )

    return jsonify({
        "campers": [c.to_dict() for c in paginated.items],
        "total": paginated.total,
        "pages": paginated.pages,
        "page": page,
    }), 200

@campers_bp.route("/<int:camper_id>", methods=["GET"])
@jwt_required()
@require_page_permission("campers", "read")
def get_camper(camper_id):
    camper = Camper.query.get_or_404(camper_id)
    return jsonify({"camper": camper.to_dict()}), 200

@campers_bp.route("/", methods=["POST"])
@jwt_required()
@require_page_permission("campers", "edit")
def create_camper():
    data = request.get_json()
    data = {k: (None if v == "" else v) for k, v in data.items()}
    if not data.get("first_name") or not data.get("last_name"):
        return jsonify({"error": "First and last name are required"}), 400

    import re
    first_name = re.sub(r'<[^>]*>', '', str(data["first_name"])).strip()[:100]
    last_name = re.sub(r'<[^>]*>', '', str(data["last_name"])).strip()[:100]

    if not first_name or not last_name:
        return jsonify({"error": "Invalid attendee name provided"}), 400

    family_group = data.get("family_group")
    waiver_status = data.get("waiver_submitted", False)
    if family_group and not waiver_status:
        existing_family_member = Camper.query.filter_by(family_group=family_group, waiver_submitted=True).first()
        if existing_family_member:
            waiver_status = True

    kayaking = 0
    if data.get("kayaking") is not None:
        try:
            kayaking = int(data.get("kayaking"))
        except (ValueError, TypeError):
            pass

    boat_tour = 0
    if data.get("boat_tour") is not None:
        try:
            boat_tour = int(data.get("boat_tour"))
        except (ValueError, TypeError):
            pass

    camper = Camper(
        first_name=first_name,
        last_name=last_name,
        date_of_birth=data.get("date_of_birth"),
        age=data.get("age"),
        gender=data.get("gender"),
        cabin_group=data.get("cabin_group"),
        family_group=family_group,
        guardian_name=data.get("guardian_name"),
        guardian_phone=data.get("guardian_phone"),
        allergies=data.get("allergies"),
        waiver_submitted=waiver_status,
        registration_status=data.get("registration_status", "registered"),
        notes=data.get("notes"),
        kayaking=kayaking,
        boat_tour=boat_tour,
        team_name=data.get("team_name"),
    )
    db.session.add(camper)
    db.session.commit()
    from utils.logging import log_action
    log_action("REGISTER_CAMPER", f"Registered camper {camper.first_name} {camper.last_name} (ID: {camper.id})")
    return jsonify({"camper": camper.to_dict()}), 201

MAX_ATTENDEES_PER_SIGNUP = 15

def _strip_html(value, max_len=255):
    """Strip HTML tags and truncate free-text input before it hits the database."""
    import re
    if value is None:
        return None
    cleaned = re.sub(r'<[^>]*>', '', str(value)).strip()
    return cleaned[:max_len] if cleaned else None

@campers_bp.route("/public-signup", methods=["POST"])
@rate_limit(5, 60)
def public_signup():
    from models import Setting
    reg_status = Setting.query.filter_by(key="registration_status").first()
    if reg_status and reg_status.value:
        status = reg_status.value.lower()
        if status == "not_open":
            return jsonify({"error": "Registration is not open yet."}), 403
        elif status == "closed":
            return jsonify({"error": "Registration is currently closed."}), 403
            
    reg_closed = Setting.query.filter_by(key="registration_closed").first()
    if reg_closed and reg_closed.value.lower() == "true":
        return jsonify({"error": "Registration is currently closed."}), 403

    data = request.get_json()
    if not data:
        return jsonify({"error": "No registration details provided"}), 400

    # Honeypot anti-spam check: this field is hidden from real users via CSS
    # and should always arrive empty. Bots that auto-fill every form field
    # will populate it, so we pretend to succeed without writing anything.
    if data.get("website"):
        return jsonify({
            "message": "Registration successful!",
            "campers": []
        }), 201

    phone = _strip_html(data.get("phone"), max_len=30)
    email = _strip_html(data.get("email"), max_len=255)
    attendees = data.get("attendees", [])

    if not phone:
        return jsonify({"error": "Guardian phone number is required"}), 400

    if not attendees:
        return jsonify({"error": "At least one attendee is required"}), 400

    if not isinstance(attendees, list) or len(attendees) > MAX_ATTENDEES_PER_SIGNUP:
        return jsonify({"error": f"A single registration can include at most {MAX_ATTENDEES_PER_SIGNUP} attendees. Please contact camp staff for larger groups."}), 400

    # Serialize family-group number assignment to avoid a race condition where two
    # concurrent signups could compute the same "next" family group number.
    is_mysql = "mysql" in str(db.engine.url)
    if is_mysql:
        db.session.execute(db.text("SELECT GET_LOCK('public_signup_family_group', 10)"))

    try:
        # Auto assign family group starting from 1001
        all_family_groups = db.session.query(Camper.family_group).filter(Camper.family_group.isnot(None)).distinct().all()
        max_num = 1000
        for (fg,) in all_family_groups:
            if fg and fg.isdigit():
                max_num = max(max_num, int(fg))
        family_group = str(max_num + 1)

        created_campers = []

        waiver_status = False
        existing_family_member = Camper.query.filter_by(family_group=family_group, waiver_submitted=True).first()
        if existing_family_member:
            waiver_status = True

        for att in attendees:
            first_name = att.get("first_name")
            last_name = att.get("last_name")
            age = att.get("age")
            gender = att.get("gender")
            allergies = _strip_html(att.get("allergies"), max_len=500)
            tshirt_size = _strip_html(att.get("tshirt_size"), max_len=20)
            indian_size = _strip_html(att.get("indian_size"), max_len=20)

            kayaking = 0
            if att.get("kayaking") is not None:
                try:
                    kayaking = int(att.get("kayaking"))
                except (ValueError, TypeError):
                    pass

            boat_tour = 0
            if att.get("boat_tour") is not None:
                try:
                    boat_tour = int(att.get("boat_tour"))
                except (ValueError, TypeError):
                    pass

            if not first_name or not last_name:
                return jsonify({"error": "First and last name are required for all attendees"}), 400

            # Strip html tags and truncate name strings to prevent XSS and DB truncation errors
            first_name = _strip_html(first_name, max_len=100)
            last_name = _strip_html(last_name, max_len=100)

            if not first_name or not last_name:
                return jsonify({"error": "Invalid attendee name provided"}), 400

            parsed_age = None
            if age is not None and str(age).strip() != "":
                try:
                    parsed_age = int(age)
                except ValueError:
                    return jsonify({"error": f"Invalid age for attendee {first_name}"}), 400

                if parsed_age < 18 and (parsed_age is None or parsed_age < 0):
                    return jsonify({"error": f"Valid age is required for child attendee {first_name}"}), 400

            camper = Camper(
                first_name=first_name,
                last_name=last_name,
                age=parsed_age,
                gender=gender if gender in ["male", "female"] else None,
                family_group=str(family_group),
                guardian_name="Self" if parsed_age is None or parsed_age >= 18 else None,
                guardian_phone=phone,
                allergies=allergies,
                waiver_submitted=waiver_status,
                registration_status="registered",
                notes=f"Public Signup. Email: {email or 'N/A'}",
                kayaking=kayaking,
                boat_tour=boat_tour
            )
            db.session.add(camper)
            db.session.flush()

            if tshirt_size or indian_size:
                tshirt = Tshirt(
                    camper_id=camper.id,
                    attendee_name=f"{first_name} {last_name}",
                    tshirt_size=tshirt_size or "Adult M",
                    indian_size=indian_size
                )
                db.session.add(tshirt)

            created_campers.append(camper)

        db.session.commit()
    finally:
        if is_mysql:
            db.session.execute(db.text("SELECT RELEASE_LOCK('public_signup_family_group')"))

    from utils.logging import log_action
    log_action("PUBLIC_SIGNUP", f"Public signup for family group {family_group} with {len(created_campers)} attendees")

    return jsonify({
        "message": "Registration successful!",
        "campers": [c.to_dict() for c in created_campers]
    }), 201

@campers_bp.route("/<int:camper_id>", methods=["PUT"])
@jwt_required()
def update_camper(camper_id):
    claims = get_jwt()
    role = claims.get("role", "user")

    has_campers_edit = check_permission(role, "campers", "edit")
    has_apparel_edit = check_permission(role, "apparel", "edit")

    if not has_campers_edit and not has_apparel_edit:
        return jsonify({"error": "Access denied"}), 403

    camper = Camper.query.get_or_404(camper_id)
    data = request.get_json() or {}
    data = {k: (None if v == "" else v) for k, v in data.items()}

    # Limit updates to only t-shirt size fields if they lack full campers edit permission
    if not has_campers_edit and has_apparel_edit:
        allowed_keys = {"tshirt_size", "indian_size"}
        data = {k: v for k, v in data.items() if k in allowed_keys}

    # Camp Director can only modify outdoor activities
    if role == "director":
        for field in ["kayaking", "boat_tour"]:
            if field in data:
                try:
                    val = int(data[field]) if data[field] is not None else 0
                except (ValueError, TypeError):
                    val = 0
                setattr(camper, field, val)
    else:
        fields = [
            "first_name", "last_name", "date_of_birth", "age", "gender", "grade",
            "cabin_group", "session", "family_group", "guardian_name", "guardian_phone", "guardian_email",
            "emergency_contact", "emergency_phone", "allergies", "medical_notes",
            "medications", "registration_status", "payment_status", "notes", "waiver_submitted",
            "kayaking", "boat_tour", "team_name"
        ]
        
        waiver_changed = "waiver_submitted" in data and data["waiver_submitted"] != camper.waiver_submitted

        from models.permission import PagePermission
        from routes.permissions import DEFAULT_PERMISSIONS

        has_teams_edit = (DEFAULT_PERMISSIONS.get(role, {}).get("teams") == "edit")
        custom_teams_perm = PagePermission.query.filter_by(role=role, page_key="teams").first()
        if custom_teams_perm:
            has_teams_edit = (custom_teams_perm.access_level == "edit")

        for field in fields:
            if field in data:
                if field == "team_name" and not has_teams_edit:
                    continue
                val = data[field]
                if field in ["first_name", "last_name"] and val is not None:
                    import re
                    val = re.sub(r'<[^>]*>', '', str(val)).strip()[:100]
                if field in ["kayaking", "boat_tour"]:
                    try:
                        val = int(val) if val is not None else 0
                    except (ValueError, TypeError):
                        val = 0
                setattr(camper, field, val)

        if "tshirt_size" in data or "indian_size" in data:
            tshirt = Tshirt.query.filter_by(camper_id=camper.id).first()
            t_size = data["tshirt_size"] if "tshirt_size" in data else (tshirt.tshirt_size if tshirt else "")
            ind_size = data["indian_size"] if "indian_size" in data else (tshirt.indian_size if tshirt else None)

            if t_size or ind_size:
                if tshirt:
                    tshirt.tshirt_size = t_size or "Adult M"
                    tshirt.indian_size = ind_size
                    tshirt.attendee_name = f"{camper.first_name} {camper.last_name}"
                else:
                    tshirt = Tshirt(
                        camper_id=camper.id,
                        attendee_name=f"{camper.first_name} {camper.last_name}",
                        tshirt_size=t_size or "Adult M",
                        indian_size=ind_size
                    )
                    db.session.add(tshirt)
            else:
                if tshirt:
                    db.session.delete(tshirt)

        if waiver_changed and camper.family_group:
            Camper.query.filter_by(family_group=camper.family_group).update({"waiver_submitted": camper.waiver_submitted})
            from utils.logging import log_action
            log_action("UPDATE_FAMILY_WAIVER", f"Synchronized waiver submission status ({camper.waiver_submitted}) for Family Group #{camper.family_group}")

    db.session.commit()
    from utils.logging import log_action
    log_action("UPDATE_CAMPER", f"Updated camper {camper.first_name} {camper.last_name} (ID: {camper.id})")
    return jsonify({"camper": camper.to_dict()}), 200

@campers_bp.route("/<int:camper_id>", methods=["DELETE"])
@jwt_required()
@require_page_permission("campers", "edit")
def delete_camper(camper_id):
    if not require_admin():
        return jsonify({"error": "Admin access required"}), 403

    camper = Camper.query.get_or_404(camper_id)
    camper_name = f"{camper.first_name} {camper.last_name}"

    # Auto re-assign Head of Family if deleted camper was the designated head
    if camper.family_group:
        from models import FamilyPayment
        pay_record = FamilyPayment.query.filter_by(family_group=camper.family_group).first()
        if pay_record and pay_record.head_camper_id == camper.id:
            next_head = Camper.query.filter(Camper.family_group == camper.family_group, Camper.id != camper.id).first()
            pay_record.head_camper_id = next_head.id if next_head else None

    db.session.delete(camper)
    db.session.commit()
    from utils.logging import log_action
    log_action("DELETE_CAMPER", f"Deleted camper {camper_name} (ID: {camper_id})")
    return jsonify({"message": "Camper deleted"}), 200

@campers_bp.route("/stats", methods=["GET"])
@jwt_required()
def get_stats():
    total = Camper.query.count()
    registered = Camper.query.filter_by(registration_status="registered").count()
    checked_in = sum(
        1 for c in Camper.query.all()
        if any(ci.checked_out_at is None for ci in c.checkins)
    )
    waivers_submitted = Camper.query.filter_by(waiver_submitted=True).count()
    total_families = db.session.query(Camper.family_group).filter(
        Camper.family_group.isnot(None),
        Camper.family_group != ""
    ).distinct().count()

    return jsonify({
        "total_registered": total,
        "status_registered": registered,
        "checked_in": checked_in,
        "waivers_submitted": waivers_submitted,
        "total_families": total_families,
    }), 200

def parse_custom_activities(notes_str):
    import re, json
    if not notes_str:
        return {}
    match = re.search(r'<!-- ACTIVITIES_JSON:\s*(.*?)\s*-->', notes_str)
    if match:
        try:
            return json.loads(match.group(1))
        except Exception:
            pass
    return {}

@campers_bp.route("/outdoor", methods=["GET"])
@jwt_required()
def get_outdoor_activities():
    raw_campers = Camper.query.filter(
        (Camper.kayaking > 0) | 
        (Camper.boat_tour > 0) |
        (Camper.notes.like('%<!-- ACTIVITIES_JSON:%'))
    ).all()
    
    campers = []
    for c in raw_campers:
        custom_acts = parse_custom_activities(c.notes)
        total_spots = (c.kayaking or 0) + (c.boat_tour or 0) + sum(int(v) for v in custom_acts.values())
        if total_spots > 0:
            campers.append(c)
    
    total_kayaking = sum(c.kayaking for c in campers)
    total_boat_tour = sum(c.boat_tour for c in campers)
    
    return jsonify({
        "campers": [c.to_dict() for c in campers],
        "total_kayaking": total_kayaking,
        "total_boat_tour": total_boat_tour
    }), 200

@campers_bp.route("/export-excel", methods=["GET"])
@jwt_required()
def export_campers_excel():
    claims = get_jwt()
    role = claims.get("role", "user")
    if not check_permission(role, "campers", "read") and not check_permission(role, "teams", "read") and not check_permission(role, "apparel", "read"):
        return jsonify({"error": "Access denied"}), 403

    search = request.args.get("search", "").strip()
    status = request.args.get("status", "").strip()

    query = Camper.query

    if search:
        like = f"%{search}%"
        matched_families = db.session.query(Camper.family_group).filter(
            db.or_(
                Camper.first_name.ilike(like),
                Camper.last_name.ilike(like),
                Camper.guardian_name.ilike(like),
                Camper.family_group.ilike(like),
            )
        ).filter(Camper.family_group.isnot(None), Camper.family_group != '').distinct().all()
        
        family_ids = [f[0] for f in matched_families if f[0]]
        
        if family_ids:
            query = query.filter(
                db.or_(
                    Camper.first_name.ilike(like),
                    Camper.last_name.ilike(like),
                    Camper.guardian_name.ilike(like),
                    Camper.family_group.ilike(like),
                    Camper.family_group.in_(family_ids)
                )
            )
        else:
            query = query.filter(
                db.or_(
                    Camper.first_name.ilike(like),
                    Camper.last_name.ilike(like),
                    Camper.guardian_name.ilike(like),
                    Camper.family_group.ilike(like),
                )
            )
    if status:
        query = query.filter(Camper.registration_status == status)

    order_by_clause = db.case(
        (db.or_(Camper.family_group.is_(None), Camper.family_group == ''), 1),
        else_=0
    )
    ordered_query = query.order_by(
        order_by_clause,
        Camper.family_group,
        Camper.last_name,
        Camper.first_name
    )

    items = ordered_query.all()

    import openpyxl
    from io import BytesIO
    from flask import send_file

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Campers"

    ws.append(["Camper Name"])

    for c in items:
        ws.append([f"{c.first_name} {c.last_name}"])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 15)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    from utils.logging import log_action
    log_action("EXPORT_CAMPERS_EXCEL", f"Exported {len(items)} camper names to Excel")

    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="gca_campers_names.xlsx"
    )

@campers_bp.route("/cabins-pdf", methods=["GET"])
@jwt_required()
@require_page_permission("cabins", "edit")
def download_cabins_pdf():
    from io import BytesIO
    from flask import send_file
    import datetime
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    # 1. Fetch all registered campers
    campers = Camper.query.filter_by(registration_status="registered").all()
    
    from utils.logging import log_action
    log_action("PRINT_CABINS_REPORT", f"Downloaded/printed camp cabin room assignments report containing {len(campers)} registered campers")
    
    # 2. Group campers by cabin and room
    cabins_map = {}
    unassigned = []
    
    for c in campers:
        if c.cabin_group:
            parts = c.cabin_group.split(" | ")
            if len(parts) == 2:
                cab_name = parts[0].strip()
                room_name = parts[1].strip()
            else:
                cab_name = c.cabin_group.strip()
                room_name = "General"
                
            if cab_name not in cabins_map:
                cabins_map[cab_name] = {}
            if room_name not in cabins_map[cab_name]:
                cabins_map[cab_name][room_name] = []
            cabins_map[cab_name][room_name].append(c)
        else:
            unassigned.append(c)
            
    sorted_cabins = sorted(cabins_map.keys(), key=lambda x: x.lower())
    
    import datetime
    current_year = datetime.date.today().year
    camp_title = f"GCA {current_year} Church Camp"

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    doc.title = f"{camp_title} - Cabin Assignments"
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        leading=24,
        textColor=colors.HexColor('#1E4D2B'),
        alignment=1,
        spaceAfter=10
    )
    subtitle_style = ParagraphStyle(
        'DocSub',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#6B7280'),
        alignment=1,
        spaceAfter=20
    )
    cabin_hdr_style = ParagraphStyle(
        'CabinHeader',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=13,
        leading=16,
        textColor=colors.HexColor('#1E4D2B'),
        spaceBefore=14,
        spaceAfter=6,
        keepWithNext=True
    )
    room_hdr_style = ParagraphStyle(
        'RoomHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        leading=13,
        textColor=colors.HexColor('#C8972B'),
        spaceBefore=6,
        spaceAfter=4,
        keepWithNext=True
    )
    table_cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        leading=11
    )
    table_hdr_cell_style = ParagraphStyle(
        'TableHdrCell',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8.5,
        leading=11,
        textColor=colors.white
    )
    
    elements = []
    
    # Header
    elements.append(Paragraph("Grace Christian Assembly", title_style))
    elements.append(Paragraph(f"{camp_title} - Cabin Assignments", ParagraphStyle('DocTitleSub', parent=title_style, fontSize=13, leading=16, spaceAfter=4)))
    
    current_time = datetime.datetime.now().strftime("%B %d, %Y at %I:%M %p")
    elements.append(Paragraph(f"Generated on {current_time}", subtitle_style))
    
    # For each cabin
    for cab in sorted_cabins:
        elements.append(Paragraph(f"⛺ Cabin: {cab}", cabin_hdr_style))
        
        rooms_map = cabins_map[cab]
        sorted_rooms = sorted(rooms_map.keys(), key=lambda x: x.lower())
        
        for rm in sorted_rooms:
            elements.append(Paragraph(f"🚪 Room: {rm}", room_hdr_style))
            
            data = [[
                Paragraph("Name", table_hdr_cell_style),
                Paragraph("Guardian", table_hdr_cell_style)
            ]]
            
            room_campers = rooms_map[rm]
            room_campers.sort(key=lambda x: f"{x.first_name} {x.last_name}".lower())
            
            for c in room_campers:
                data.append([
                    Paragraph(f"{c.first_name} {c.last_name}", table_cell_style),
                    Paragraph(c.guardian_name or "-", table_cell_style)
                ])
                
            col_widths = [266, 266]
            
            t = Table(data, colWidths=col_widths)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E4D2B')),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('BOTTOMPADDING', (0,0), (-1,0), 4),
                ('TOPPADDING', (0,0), (-1,0), 4),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#EDE8DC')),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8F5EE')]),
                ('BOTTOMPADDING', (0,1), (-1,-1), 4),
                ('TOPPADDING', (0,1), (-1,-1), 4),
            ]))
            
            elements.append(t)
            elements.append(Spacer(1, 6))
            
        elements.append(Spacer(1, 10))
        
    # Unassigned
    if unassigned:
        elements.append(Paragraph("❓ Unassigned Campers", cabin_hdr_style))
        data = [[
            Paragraph("Name", table_hdr_cell_style),
            Paragraph("Guardian", table_hdr_cell_style)
        ]]
        unassigned.sort(key=lambda x: f"{x.first_name} {x.last_name}".lower())
        for c in unassigned:
            data.append([
                Paragraph(f"{c.first_name} {c.last_name}", table_cell_style),
                Paragraph(c.guardian_name or "-", table_cell_style)
            ])
            
        col_widths = [266, 266]
        t = Table(data, colWidths=col_widths)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#6B7280')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,0), 4),
            ('TOPPADDING', (0,0), (-1,0), 4),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#EDE8DC')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8F5EE')]),
            ('BOTTOMPADDING', (0,1), (-1,-1), 4),
            ('TOPPADDING', (0,1), (-1,-1), 4),
        ]))
        elements.append(t)
        
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name="gca_camp_cabins_report.pdf"
    )

@campers_bp.route("/upload-teams", methods=["POST"])
@jwt_required()
@require_page_permission("teams", "edit")
def upload_teams():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if not file or not file.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Please upload a valid Excel (.xlsx) file"}), 400

    import openpyxl
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            return jsonify({"error": "Excel sheet contains no data rows"}), 400
    except Exception as e:
        return jsonify({"error": f"Failed to parse Excel file: {str(e)}"}), 400

    header = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    name_idx = -1
    team_idx = -1

    for i, col in enumerate(header):
        if "name" in col and name_idx == -1:
            name_idx = i
        elif "team" in col and team_idx == -1:
            team_idx = i

    if name_idx == -1:
        name_idx = 1 if len(header) > 1 else 0
    if team_idx == -1:
        team_idx = 4 if len(header) > 4 else (len(header) - 1)

    campers = Camper.query.all()
    campers_by_name = {f"{c.first_name.strip()} {c.last_name.strip()}".lower(): c for c in campers}

    matched_count = 0
    updated_count = 0
    unmatched = []

    for row in rows[1:]:
        if not row or len(row) <= name_idx or not row[name_idx]:
            continue

        raw_name = str(row[name_idx]).strip()
        raw_team = str(row[team_idx]).strip() if len(row) > team_idx and row[team_idx] is not None else ""

        if not raw_name or raw_name.lower() == "name":
            continue

        canonical_team = raw_team
        lowered_team = raw_team.lower()
        if "1" in lowered_team or "peter" in lowered_team:
            canonical_team = "Team 1"
        elif "2" in lowered_team or "paul" in lowered_team:
            canonical_team = "Team 2"

        key = raw_name.lower()
        target_camper = campers_by_name.get(key)

        if not target_camper:
            parts = raw_name.split()
            if len(parts) >= 2:
                first = parts[0].lower()
                last = parts[-1].lower()
                for c_key, c_obj in campers_by_name.items():
                    if first in c_key and last in c_key:
                        target_camper = c_obj
                        break

        if target_camper:
            matched_count += 1
            if target_camper.team_name != canonical_team:
                target_camper.team_name = canonical_team
                updated_count += 1
        else:
            unmatched.append(raw_name)

    db.session.commit()

    return jsonify({
        "message": f"Successfully processed {len(rows)-1} rows. Matched {matched_count} campers.",
        "processed": len(rows) - 1,
        "matched": matched_count,
        "updated": updated_count,
        "unmatched": unmatched
    }), 200


@campers_bp.route("/upload-cabins", methods=["POST"])
@jwt_required()
@require_page_permission("cabins", "edit")
def upload_cabins_config():
    claims = get_jwt()
    if claims.get("role") != "owner":
        return jsonify({"error": "Only owners are allowed to import cabins Excel sheets."}), 403

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if not file or not file.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Please upload a valid Excel (.xlsx) file"}), 400

    import openpyxl
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            return jsonify({"error": "Excel sheet contains no data rows"}), 400
    except Exception as e:
        return jsonify({"error": f"Failed to parse Excel file: {str(e)}"}), 400

    # Dynamically find the header row containing "cabin" and "room" (case-insensitive)
    header_row_idx = 0
    for idx, r in enumerate(rows):
        row_str = " ".join([str(cell).lower() for cell in r if cell is not None])
        if "cabin" in row_str and "room" in row_str:
            header_row_idx = idx
            break

    header = [str(cell).strip().lower() if cell is not None else "" for cell in rows[header_row_idx]]
    
    cabin_idx = -1
    room_idx = -1
    occupancy_idx = -1
    location_idx = -1
    handicap_idx = -1
    
    king_idx = -1
    queen_idx = -1
    full_idx = -1
    twin_idx = -1
    bunk_idx = -1
    
    generic_bed_indices = []
    is_generic_beds = False

    for i, col in enumerate(header):
        if "cabin" in col:
            cabin_idx = i
        elif "room" in col:
            room_idx = i
        elif "occupancy" in col or "capacity" in col or "max" in col:
            occupancy_idx = i
        elif "location" in col or "floor" in col or "level" in col or "upstairs" in col or "downstairs" in col:
            location_idx = i
        elif "handicap" in col or "access" in col:
            handicap_idx = i
        elif "bed" in col:
            # Check if size is specified in column header
            if "king" in col: king_idx = i
            elif "queen" in col: queen_idx = i
            elif "full" in col: full_idx = i
            elif "twin" in col: twin_idx = i
            elif "bunk" in col: bunk_idx = i
            else:
                generic_bed_indices.append(i)
                is_generic_beds = True

    if cabin_idx == -1: cabin_idx = 0
    if room_idx == -1: room_idx = 1 if len(header) > 1 else 0

    cabins_dict = {}

    for row in rows[header_row_idx + 1:]:
        if not row or len(row) <= cabin_idx or row[cabin_idx] is None:
            continue

        cabin_name = str(row[cabin_idx]).strip()
        room_name = str(row[room_idx]).strip() if len(row) > room_idx and row[room_idx] is not None else "General"
        
        if not cabin_name or cabin_name.lower() == "cabin":
            continue

        max_occupancy = 4
        if occupancy_idx != -1 and len(row) > occupancy_idx and row[occupancy_idx] is not None:
            try:
                max_occupancy = int(row[occupancy_idx])
            except ValueError:
                pass

        location = "Single Level"
        if location_idx != -1 and len(row) > location_idx and row[location_idx] is not None:
            val = str(row[location_idx]).strip().lower()
            if "up" in val:
                location = "Upstairs"
            elif "down" in val:
                location = "Downstairs"
            else:
                location = "Single Level"

        handicap_accessible = False
        if handicap_idx != -1 and len(row) > handicap_idx and row[handicap_idx] is not None:
            val = str(row[handicap_idx]).strip().lower()
            if val in ["yes", "true", "1", "y", "t", "accessible"]:
                handicap_accessible = True

        beds = { "king": 0, "queen": 0, "full": 0, "twin": 0, "bunk": 0 }
        
        if is_generic_beds:
            for idx in generic_bed_indices:
                if idx < len(row) and row[idx] is not None:
                    val = str(row[idx]).strip().lower()
                    if "king" in val: beds["king"] += 1
                    elif "queen" in val: beds["queen"] += 1
                    elif "full" in val: beds["full"] += 1
                    elif "twin" in val: beds["twin"] += 1
                    elif "bunk" in val: beds["bunk"] += 1
        else:
            def parse_bed_qty(idx):
                if idx != -1 and len(row) > idx and row[idx] is not None:
                    try:
                        return int(row[idx])
                    except ValueError:
                        pass
                return 0

            beds["king"] = parse_bed_qty(king_idx)
            beds["queen"] = parse_bed_qty(queen_idx)
            beds["full"] = parse_bed_qty(full_idx)
            beds["twin"] = parse_bed_qty(twin_idx)
            beds["bunk"] = parse_bed_qty(bunk_idx)

        if cabin_name not in cabins_dict:
            cabins_dict[cabin_name] = {}

        cabins_dict[cabin_name][room_name] = {
            "name": room_name,
            "max_occupancy": max_occupancy,
            "location": location,
            "handicap_accessible": handicap_accessible,
            "beds": beds
        }

    config_list = []
    for cabin_name, rooms_map in cabins_dict.items():
        sorted_rooms = sorted(
            rooms_map.values(),
            key=lambda r: r["name"]
        )
        config_list.append({
            "name": cabin_name,
            "rooms": sorted_rooms
        })

    config_list.sort(key=lambda c: c["name"])

    return jsonify({
        "message": f"Successfully parsed {len(config_list)} cabins configuration",
        "config": config_list
    }), 200
