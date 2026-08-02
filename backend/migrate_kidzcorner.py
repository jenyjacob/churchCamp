"""
Migrate Kidz_Corner_Schedule.xlsx into the Kidz Corner tables (volunteers,
kids, schedule items, crafts, budget lines, AV links).

Reads sheets UP TO AND INCLUDING "AudioVideo" only:
  - "People"                -> KidzCornerVolunteer (cols A/B) + KidzCornerKid (cols D/E/F)
  - "Template of Schedule"   -> KidzCornerScheduleItem (day="Template")
  - "Budget.Expenses"        -> KidzCornerBudgetItem
  - "Friday Night"           -> KidzCornerScheduleItem (day="Friday Night") + KidzCornerCraft
  - "Saturday Morning"       -> KidzCornerScheduleItem (day="Saturday Morning") + KidzCornerCraft
  - "Saturday Night"         -> KidzCornerScheduleItem (day="Saturday Night")
  - "Sunday Morning"         -> KidzCornerScheduleItem (day="Sunday Morning") (sheet may be empty)
  - "AudioVideo"             -> KidzCornerAVLink

Any sheets after "AudioVideo" (Puppet Show sheets, Calendar) are intentionally
NOT migrated, per request.

Usage (run from the backend/ directory, same folder as app.py):
    python3 migrate_kidzcorner.py /path/to/Kidz_Corner_Schedule.xlsx
    python3 migrate_kidzcorner.py /path/to/Kidz_Corner_Schedule.xlsx --clear

--clear wipes any existing Kidz Corner rows before importing (use on a fresh
import / re-import; without it, running the script twice will duplicate rows).

This uses the Flask app + SQLAlchemy models directly (not the HTTP API), so it
works against whatever database your .env / config.py already points to
(SQLite locally, MySQL in production) without needing a running server or a
login token.
"""

import sys
import os
import argparse
import datetime

import openpyxl

# --- Load the Flask app & models -------------------------------------------------
sys.path.insert(0, ".")
from app import create_app
from db import db
from models import (
    KidzCornerVolunteer,
    KidzCornerKid,
    KidzCornerScheduleItem,
    KidzCornerCraft,
    KidzCornerBudgetItem,
    KidzCornerAVLink,
)

DAY_SHEETS = ["Friday Night", "Saturday Morning", "Saturday Night", "Sunday Morning"]
REQUIRED_SHEETS = ["People", "Template of Schedule", "Budget.Expenses"] + DAY_SHEETS + ["AudioVideo"]


def clean(value):
    """Strip strings, turn blank/placeholder values into None."""
    if value is None:
        return None
    text = str(value).strip()
    if text == "" or text == "—":
        return None
    return text


def clean_int(value):
    text = clean(value)
    if text is None:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def clean_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean(value)
    if text is None:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def fmt_date(value):
    if value is None:
        return None
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.strftime("%Y-%m-%d")
    return clean(value)


def after_colon(text):
    """Given 'Materials: small bags, ribbon' return 'small bags, ribbon'."""
    idx = text.find(":")
    if idx == -1:
        return text.strip()
    return text[idx + 1:].strip()


def skip_to_data(rows, header_markers):
    """
    Return the rows following the column-header row (matched by its first
    cell, case-insensitively, against header_markers). Sheets in this
    workbook have an inconsistent number of blank rows between the title
    and the header, so we scan for the header rather than assuming a fixed
    row index.
    """
    for i, row in enumerate(rows):
        c0 = row[0] if len(row) > 0 else None
        if c0 and str(c0).strip().lower() in header_markers:
            return rows[i + 1:]
    return rows


# ---------------------------------------------------------------------------
# People sheet -> volunteers + kids
# ---------------------------------------------------------------------------

def parse_people(ws):
    volunteers = []
    kids = []
    rows = list(ws.iter_rows(values_only=True))
    v_order = 0
    k_order = 0
    for row in rows[2:]:  # skip group-header row + column-header row
        row = list(row) + [None] * (6 - len(row))
        v_name, v_assignment = row[0], row[1]
        k_name, k_age, k_allergies = row[3], row[4], row[5]

        if clean(v_name):
            volunteers.append(KidzCornerVolunteer(
                name=clean(v_name),
                assignment=clean(v_assignment),
                order_index=v_order,
            ))
            v_order += 1

        if clean(k_name):
            kids.append(KidzCornerKid(
                name=clean(k_name),
                age=clean_int(k_age),
                allergies=clean(k_allergies),
                order_index=k_order,
            ))
            k_order += 1

    return volunteers, kids


# ---------------------------------------------------------------------------
# Template of Schedule sheet -> schedule items (day="Template")
# ---------------------------------------------------------------------------

def parse_template_schedule(ws):
    items = []
    rows = list(ws.iter_rows(values_only=True))
    data_rows = skip_to_data(rows, {"date"})
    order_index = 0
    for row in data_rows:
        row = list(row) + [None] * (6 - len(row))
        date_val, time_val, activity, vol_needed, items_needed, notes = row[:6]
        if not clean(activity):
            continue
        items.append(KidzCornerScheduleItem(
            day="Template",
            date=fmt_date(date_val),
            time=clean(time_val),
            activity=clean(activity),
            volunteers_needed=clean(vol_needed),
            items_needed=clean(items_needed),
            notes=clean(notes),
            order_index=order_index,
        ))
        order_index += 1
    return items


# ---------------------------------------------------------------------------
# Budget.Expenses sheet -> budget line items
# ---------------------------------------------------------------------------

def parse_budget(ws):
    items = []
    rows = list(ws.iter_rows(values_only=True))
    order_index = 0
    for row in rows[1:]:  # skip column-header row
        row = list(row) + [None] * (6 - len(row))
        month, income_actual, expenses_actual, expenses_projected, related_files, notes = row[:6]

        # Skip the trailing totals row: it has no month and no notes, only sums.
        if not clean(month) and not clean(notes):
            continue

        if not any([clean(month), clean_float(income_actual), clean_float(expenses_actual),
                    clean_float(expenses_projected), clean(related_files), clean(notes)]):
            continue

        items.append(KidzCornerBudgetItem(
            month=clean(month),
            income_actual=clean_float(income_actual),
            expenses_actual=clean_float(expenses_actual),
            expenses_projected=clean_float(expenses_projected),
            related_files=clean(related_files),
            notes=clean(notes),
            order_index=order_index,
        ))
        order_index += 1
    return items


# ---------------------------------------------------------------------------
# Day sheets (Friday Night / Saturday Morning / Saturday Night / Sunday Morning)
# -> schedule items + an optional trailing craft block
# ---------------------------------------------------------------------------

def parse_day_sheet(ws, day_name):
    items = []
    crafts = []
    rows = list(ws.iter_rows(values_only=True))
    data_rows = skip_to_data(rows, {"time"})

    order_index = 0
    craft = None

    for row in data_rows:
        row = list(row) + [None] * (5 - len(row))
        c0, c1, c2, c3, c4 = row[:5]

        if not any(clean(c) for c in (c0, c1, c2, c3, c4)):
            continue  # blank row

        first_text = clean(c0)

        if first_text and first_text.lower().startswith("craft"):
            if craft is not None:
                crafts.append(craft)
            craft = KidzCornerCraft(
                day=day_name,
                title=after_colon(first_text) or first_text,
                order_index=len(crafts),
            )
            continue

        if craft is not None and first_text:
            low = first_text.lower()
            if low.startswith("materials"):
                craft.materials = after_colon(first_text)
            elif low.startswith("how-to") or low.startswith("how to"):
                craft.how_to = after_colon(first_text)
            elif low.startswith("ages"):
                craft.ages = after_colon(first_text)
            elif low.startswith("things to bring"):
                craft.things_to_bring = after_colon(first_text)
            continue

        # A normal schedule data row: Time, Activity, Volunteers Needed, Items Needed, Notes
        time_val, activity, vol_needed, items_needed, notes = c0, c1, c2, c3, c4
        if not clean(activity):
            continue

        items.append(KidzCornerScheduleItem(
            day=day_name,
            time=clean(time_val),
            activity=clean(activity),
            volunteers_needed=clean(vol_needed),
            items_needed=clean(items_needed),
            notes=clean(notes),
            order_index=order_index,
        ))
        order_index += 1

    if craft is not None:
        crafts.append(craft)

    return items, crafts


# ---------------------------------------------------------------------------
# AudioVideo sheet -> AV links, grouped under whichever header row precedes them
# ---------------------------------------------------------------------------

def parse_audiovideo(ws):
    links = []
    current_category = "Action Song List"
    order_index = 0
    for row in ws.iter_rows(values_only=True):
        row = list(row) + [None] * (2 - len(row))
        label, url = row[:2]
        label = clean(label)
        url = clean(url)
        if not label:
            continue
        if not url:
            # A row with only a label and no URL is a category header, e.g.
            # "Action Song List:" or "Dance Party".
            current_category = label.rstrip(":").strip()
            continue
        links.append(KidzCornerAVLink(
            category=current_category,
            label=label,
            url=url,
            order_index=order_index,
        ))
        order_index += 1
    return links


def migrate(xlsx_path, clear=False):
    # 1. Load the workbook FIRST so a bad file never touches the database.
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        print(f"Error: could not open '{xlsx_path}': {e}")
        return

    missing = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
    if missing:
        print(f"Error: workbook is missing expected sheet(s): {missing}. Aborting.")
        return

    volunteers, kids = parse_people(wb["People"])
    template_items = parse_template_schedule(wb["Template of Schedule"])
    budget_items = parse_budget(wb["Budget.Expenses"])

    schedule_items = list(template_items)
    crafts = []
    for day_name in DAY_SHEETS:
        day_items, day_crafts = parse_day_sheet(wb[day_name], day_name)
        schedule_items.extend(day_items)
        crafts.extend(day_crafts)

    av_links = parse_audiovideo(wb["AudioVideo"])

    print(
        f"Parsed from workbook: {len(volunteers)} volunteers, {len(kids)} kids, "
        f"{len(schedule_items)} schedule items, {len(crafts)} crafts, "
        f"{len(budget_items)} budget lines, {len(av_links)} AV links."
    )

    if not any([volunteers, kids, schedule_items, crafts, budget_items, av_links]):
        print("Nothing parsed from the workbook — aborting without touching the database.")
        return

    # 2. Only now touch the database.
    app = create_app()
    with app.app_context():
        db.create_all()

        if clear:
            deleted = (
                KidzCornerVolunteer.query.delete(),
                KidzCornerKid.query.delete(),
                KidzCornerScheduleItem.query.delete(),
                KidzCornerCraft.query.delete(),
                KidzCornerBudgetItem.query.delete(),
                KidzCornerAVLink.query.delete(),
            )
            db.session.commit()
            print(
                f"Cleared existing rows: {deleted[0]} volunteers, {deleted[1]} kids, "
                f"{deleted[2]} schedule items, {deleted[3]} crafts, "
                f"{deleted[4]} budget lines, {deleted[5]} AV links."
            )

        db.session.add_all(volunteers)
        db.session.add_all(kids)
        db.session.add_all(schedule_items)
        db.session.add_all(crafts)
        db.session.add_all(budget_items)
        db.session.add_all(av_links)
        db.session.commit()

    print("Migration complete.")
    print(f"  Volunteers:     {len(volunteers)}")
    print(f"  Kids:           {len(kids)}")
    print(f"  Schedule items: {len(schedule_items)}")
    print(f"  Crafts:         {len(crafts)}")
    print(f"  Budget lines:   {len(budget_items)}")
    print(f"  AV links:       {len(av_links)}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Migrate the Kidz Corner Excel workbook into the app's database.")
    parser.add_argument("xlsx_path", help="Path to Kidz_Corner_Schedule.xlsx")
    parser.add_argument("--clear", action="store_true",
                         help="Delete existing Kidz Corner rows before importing (use for a clean re-import).")
    args = parser.parse_args()
    xlsx_abspath = os.path.abspath(args.xlsx_path)
    migrate(xlsx_abspath, clear=args.clear)
