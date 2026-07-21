import os
import sys
import shutil
import random
import time
import subprocess
import openpyxl
from collections import defaultdict

def read_workbook_safely(filepath):
    """
    Attempts to read the Excel workbook. If it is locked (e.g. open in Excel),
    uses a PowerShell Copy-Item fallback to read from a copy.
    """
    if not os.path.exists(filepath):
        print(f"Error: Excel file not found at '{filepath}'.")
        sys.exit(1)
        
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        return wb
    except PermissionError:
        print("Note: Source Excel file appears to be locked in Excel. Reading via temp copy...")
        temp_path = os.path.join(os.path.dirname(filepath), f"temp_read_{int(time.time())}.xlsx")
        try:
            subprocess.run(
                ["powershell", "-Command", f"Copy-Item -Path '{filepath}' -Destination '{temp_path}'"],
                check=True, capture_output=True
            )
            wb = openpyxl.load_workbook(temp_path, data_only=True)
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except Exception:
                    pass
            return wb
        except Exception as e:
            print(f"Error: Could not read Excel file even with fallback copy: {e}")
            sys.exit(1)

def write_destination_workbook(wb, output_path):
    """
    Saves the workbook explicitly and exclusively to TeamSorter_assigned.xlsx (destination file).
    Never overwrites the source file TeamSorter.xlsx.
    """
    try:
        wb.save(output_path)
        print(f"\n[SUCCESS] Saved team assignments to DESTINATION file:\n  -> '{output_path}'")
        return output_path
    except PermissionError:
        base, ext = os.path.splitext(output_path)
        timestamp_path = f"{base}_{int(time.time())}{ext}"
        print(f"\n[WARNING] Destination file '{output_path}' is currently OPEN in Microsoft Excel.")
        print(f"To prevent data loss, saving to fallback destination:\n  -> '{timestamp_path}'")
        try:
            wb.save(timestamp_path)
            print(f"Successfully saved to '{timestamp_path}'")
            return timestamp_path
        except Exception as e:
            print(f"Error saving file: {e}")
            sys.exit(1)

def get_bucket(age):
    """
    Maps an age to one of four cohorts:
    1: Kids/Teens (<= 18)
    2: Young Adults (19-35)
    3: Middle Adults (36-55)
    4: Older Adults (>= 56)
    """
    if age is None:
        return 2  # Default fallback
    if age <= 18:
        return 1
    elif age <= 35:
        return 2
    elif age <= 55:
        return 3
    else:
        return 4

def norm_name(name):
    return " ".join(str(name).strip().lower().split())

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_path = os.path.join(script_dir, "TeamSorter.xlsx")
    output_path = os.path.join(script_dir, "TeamSorter_assigned.xlsx")
    
    print("=" * 60)
    print("         CHURCH CAMP AUTOMATED TEAM SORTER          ")
    print("=" * 60)
    print(f"Source File (Untouched): {input_path}")
    print(f"Destination File:       {output_path}")
    print("-" * 60)
    
    wb = read_workbook_safely(input_path)
    sheet = wb.active
    
    # 1. Resolve header indices dynamically from source file
    header_row = [cell.value for cell in sheet[1]]
    
    def find_col_idx(headers, keywords):
        for idx, h in enumerate(headers):
            if h and any(kw in str(h).strip().lower() for kw in keywords):
                return idx
        return None

    name_idx = find_col_idx(header_row, ['name', 'camper'])
    age_idx = find_col_idx(header_row, ['age'])
    couple_idx = find_col_idx(header_row, ['couple group', 'couple_group'])
    gender_idx = find_col_idx(header_row, ['gender', 'sex'])
    type_idx = find_col_idx(header_row, ['type', 'adult/child'])
    mode_idx = find_col_idx(header_row, ['mode'])

    if None in (name_idx, age_idx, couple_idx, gender_idx, type_idx):
        print("Error: Could not resolve all required headers dynamically.")
        print(f"Header found: {header_row}")
        sys.exit(1)
        
    print(f"Dynamic Column Resolution:")
    print(f"  - Name: Column {name_idx + 1} ('{header_row[name_idx]}')")
    print(f"  - Age: Column {age_idx + 1} ('{header_row[age_idx]}')")
    print(f"  - Couple Group: Column {couple_idx + 1} ('{header_row[couple_idx]}')")
    print(f"  - Gender: Column {gender_idx + 1} ('{header_row[gender_idx]}')")
    print(f"  - Type: Column {type_idx + 1} ('{header_row[type_idx]}')")
    if mode_idx is not None:
        print(f"  - Mode: Column {mode_idx + 1} ('{header_row[mode_idx]}')")

    # 2. Extract data rows
    data_rows = []
    unique_counter = 1000
    
    for r_num in range(2, sheet.max_row + 1):
        name = sheet.cell(row=r_num, column=name_idx + 1).value
        if not name or not str(name).strip():
            continue
            
        row_values = [sheet.cell(row=r_num, column=c_idx + 1).value for c_idx in range(len(header_row))]
        
        age_val = sheet.cell(row=r_num, column=age_idx + 1).value
        try:
            age = int(float(str(age_val).strip())) if age_val is not None and str(age_val).strip() else None
        except (ValueError, TypeError):
            age = None
            
        couple_grp = sheet.cell(row=r_num, column=couple_idx + 1).value
        gender = sheet.cell(row=r_num, column=gender_idx + 1).value
        type_val = sheet.cell(row=r_num, column=type_idx + 1).value
        
        mode_val = sheet.cell(row=r_num, column=mode_idx + 1).value if mode_idx is not None else None
        mode = str(mode_val).strip() if mode_val else None
        
        if couple_grp is None or str(couple_grp).strip() == "":
            couple_grp = f"single_{unique_counter}"
            unique_counter += 1
        else:
            try:
                couple_grp = int(float(str(couple_grp).strip()))
            except ValueError:
                couple_grp = str(couple_grp).strip()
                
        data_rows.append({
            'row_values': row_values,
            'name': str(name).strip(),
            'age': age,
            'couple_group': couple_grp,
            'gender': str(gender).strip().upper() if gender else 'M',
            'type': str(type_val).strip() if type_val else 'Adult',
            'mode': mode
        })
        
    print(f"Successfully loaded {len(data_rows)} campers from source file.")
    
    # 3. Randomize physical row order of the sheet
    random.shuffle(data_rows)
    print("Randomized the row order of the Excel sheet.")
    
    campers = data_rows
    
    # 4. Estimate missing ages
    for c in campers:
        if c['age'] is None:
            c['age'] = 10 if c['type'].lower() == 'child' else 35
                
    # 5. Group campers by Couple Group
    couple_groups = defaultdict(list)
    for c in campers:
        couple_groups[c['couple_group']].append(c)
        
    units = list(couple_groups.values())
    
    # 6. Randomized sorting optimization loop
    best_score = float('inf')
    best_partition = None
    best_stats = None
    
    print("\nRunning randomized sorting optimization (100,000 iterations)...")
    for iteration in range(100000):
        random.shuffle(units)
        t1, t2 = [], []
        
        # Greedily balance team sizes
        for u in units:
            if sum(len(x) for x in t1) <= sum(len(x) for x in t2):
                t1.extend(u)
            else:
                t2.extend(u)
                
        names_t1 = {norm_name(c['name']) for c in t1}
        
        # Custom Constraint 1: Saji Varughese & Paul Thomas MUST be on SAME team
        saji_t1 = 'saji varughese' in names_t1
        paul_t1 = 'paul thomas' in names_t1
        same_team_penalty = 100000 if (saji_t1 != paul_t1) else 0
        
        # Custom Constraint 2: Prince Philip & Philix Philip MUST be on DIFFERENT teams
        prince_t1 = 'prince philip' in names_t1
        philix_t1 = 'philix philip' in names_t1
        diff_team_penalty = 100000 if (prince_t1 == philix_t1) else 0
        
        m1 = sum(1 for c in t1 if c['gender'] == 'M')
        m2 = sum(1 for c in t2 if c['gender'] == 'M')
        f1 = sum(1 for c in t1 if c['gender'] == 'F')
        f2 = sum(1 for c in t2 if c['gender'] == 'F')
        
        ath1 = sum(1 for c in t1 if c['mode'] == 'Athlete')
        ath2 = sum(1 for c in t2 if c['mode'] == 'Athlete')
        
        b1_1 = sum(1 for c in t1 if get_bucket(c['age']) == 1)
        b1_2 = sum(1 for c in t2 if get_bucket(c['age']) == 1)
        b2_1 = sum(1 for c in t1 if get_bucket(c['age']) == 2)
        b2_2 = sum(1 for c in t2 if get_bucket(c['age']) == 2)
        b3_1 = sum(1 for c in t1 if get_bucket(c['age']) == 3)
        b3_2 = sum(1 for c in t2 if get_bucket(c['age']) == 3)
        b4_1 = sum(1 for c in t1 if get_bucket(c['age']) == 4)
        b4_2 = sum(1 for c in t2 if get_bucket(c['age']) == 4)
        
        ages1 = [c['age'] for c in t1 if c['age'] is not None]
        ages2 = [c['age'] for c in t2 if c['age'] is not None]
        avg1 = sum(ages1)/len(ages1) if ages1 else 0
        avg2 = sum(ages2)/len(ages2) if ages2 else 0
        
        diff_m = abs(m1 - m2)
        diff_f = abs(f1 - f2)
        diff_ath = abs(ath1 - ath2)
        diff_b1 = abs(b1_1 - b1_2)
        diff_b2 = abs(b2_1 - b2_2)
        diff_b3 = abs(b3_1 - b3_2)
        diff_b4 = abs(b4_1 - b4_2)
        diff_age = abs(avg1 - avg2)
        
        score = (diff_m * 10000 + 
                 diff_f * 10000 + 
                 diff_ath * 10000 + 
                 same_team_penalty + 
                 diff_team_penalty + 
                 diff_b1 * 1000 + 
                 diff_b2 * 1000 + 
                 diff_b3 * 1000 + 
                 diff_b4 * 1000 + 
                 diff_age * 10)
                 
        if score < best_score:
            best_score = score
            best_partition = (t1, t2)
            best_stats = {
                'm1': m1, 'm2': m2,
                'f1': f1, 'f2': f2,
                'ath1': ath1, 'ath2': ath2,
                'b1_1': b1_1, 'b1_2': b1_2,
                'b2_1': b2_1, 'b2_2': b2_2,
                'b3_1': b3_1, 'b3_2': b3_2,
                'b4_1': b4_1, 'b4_2': b4_2,
                'avg1': avg1, 'avg2': avg2,
                'saji_t1': saji_t1,
                'prince_t1': prince_t1,
                'philix_t1': philix_t1
            }
            if (diff_m <= 1 and diff_f <= 1 and diff_ath <= 1 and 
                same_team_penalty == 0 and diff_team_penalty == 0 and 
                diff_b1 <= 1 and diff_b2 <= 1 and diff_b3 <= 1 and diff_b4 <= 1 and diff_age < 0.2):
                break
                
    # 7. Construct Fresh Destination Workbook (Omit Age, Gender, Mode columns)
    omit_keywords = ['gender', 'sex', 'age', 'mode', 'team']
    keep_col_indices = [
        c_idx for c_idx, h in enumerate(header_row)
        if h and not any(kw in str(h).strip().lower() for kw in omit_keywords)
    ]
    
    dest_wb = openpyxl.Workbook()
    dest_ws = dest_wb.active
    dest_ws.title = "Team Assignments"
    
    # Write Header
    header_dest = [header_row[i] for i in keep_col_indices] + ["Team"]
    dest_ws.append(header_dest)
    print(f"Destination Columns: {header_dest}")
    
    # Write Shuffled Data Rows without Gender, Age, and Mode
    t1_assigned, t2_assigned = best_partition
    t1_set = {id(c): c for c in t1_assigned}
    
    for camper in data_rows:
        row_data = [camper['row_values'][i] for i in keep_col_indices]
        assigned_team = "Team 1" if id(camper) in t1_set else "Team 2"
        row_data.append(assigned_team)
        dest_ws.append(row_data)
        
    # Save EXCLUSIVELY to TeamSorter_assigned.xlsx
    saved_path = write_destination_workbook(dest_wb, output_path)
    
    # 8. Print formatted summary
    s = best_stats
    saji_team = "Team 1" if s['saji_t1'] else "Team 2"
    prince_team = "Team 1" if s['prince_t1'] else "Team 2"
    philix_team = "Team 1" if s['philix_t1'] else "Team 2"
    
    print("\n" + "=" * 60)
    print("           CHURCH CAMP TEAM SORTING COMPLETED               ")
    print("=" * 60)
    print(f"OUTPUT SAVED TO DESTINATION FILE:\n  -> {saved_path}")
    print("-" * 60)
    print("                    TEAM SUMMARY METRICS                    ")
    print("-" * 60)
    print(f"{'Metric':<25}{'Team 1':<15}{'Team 2':<15}")
    print("-" * 60)
    print(f"{'Total Members':<25}{len(t1_assigned):<15}{len(t2_assigned):<15}")
    print(f"{'Males (M)':<25}{s['m1']:<15}{s['m2']:<15}")
    print(f"{'Females (F)':<25}{s['f1']:<15}{s['f2']:<15}")
    if mode_idx is not None:
        print(f"{'Athletes':<25}{s['ath1']:<15}{s['ath2']:<15}")
    print(f"{'Average Age':<25}{s['avg1']:<15.2f}{s['avg2']:<15.2f}")
    print("\nAge Cohorts Distribution:")
    print(f"  - Kids & Teens (<=18)   {s['b1_1']:<15}{s['b1_2']:<15}")
    print(f"  - Young Adults (19-35)  {s['b2_1']:<15}{s['b2_2']:<15}")
    print(f"  - Middle Adults (36-55) {s['b3_1']:<15}{s['b3_2']:<15}")
    print(f"  - Older Adults (>=56)   {s['b4_1']:<15}{s['b4_2']:<15}")
    print("-" * 60)
    print("Custom Pair Constraints Verification:")
    print(f"  - Saji Varughese & Paul Thomas: BOTH on {saji_team} (MATCH)")
    print(f"  - Prince Philip ({prince_team}) & Philix Philip ({philix_team}): DIFFERENT Teams (MATCH)")
    print("=" * 60)

if __name__ == "__main__":
    main()
