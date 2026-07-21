import openpyxl
import pymysql
import os

file_path = "GCA Camp Report.xlsx"

def split_name(full_name):
    if not full_name:
        return "", ""
    parts = full_name.strip().split()
    if len(parts) == 1:
        return parts[0], ""
    return " ".join(parts[:-1]), parts[-1]

def migrate():
    # 1. Load Excel workbook first to prevent database wipe if file loading fails
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb["Adults_Children"] if "Adults_Children" in wb.sheetnames else wb.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Excel file has no data rows")
    except Exception as e:
        print(f"Error: Failed to load Excel workbook from {file_path}. Aborting migration to protect database entries.")
        print(f"Details: {e}")
        return

    header = rows[0]
    data_rows = rows[1:]

    # 2. Connect to local MySQL database (loading credentials dynamically from environment)
    import urllib.parse as urlparse
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(__file__), '.env'))
    load_dotenv(os.path.join(os.path.dirname(__file__), 'backend', '.env'))

    db_url = os.environ.get("DATABASE_URL")
    db_user = os.environ.get("DB_USER", "")
    db_password = os.environ.get("DB_PASSWORD", "")
    db_host = os.environ.get("DB_HOST", "localhost")
    db_port = int(os.environ.get("DB_PORT", 3306))
    db_name = os.environ.get("DB_NAME", "churchcamp")

    if db_url:
        try:
            parsed_url = db_url.replace("mysql+pymysql://", "http://")
            url = urlparse.urlparse(parsed_url)
            db_user = url.username or db_user
            db_password = url.password or db_password
            db_host = url.hostname or db_host
            db_port = url.port or db_port
            db_name = url.path.lstrip('/') or db_name
        except Exception as e:
            print(f"Warning: Failed to parse DATABASE_URL: {e}")

    conn = pymysql.connect(
        host=db_host, 
        user=db_user, 
        password=db_password, 
        port=db_port, 
        database=db_name
    )
    cursor = conn.cursor()
    
    try:
        # Clear existing campers to prevent duplicates (will not commit until the end of transaction)
        cursor.execute("DELETE FROM checkins")
        cursor.execute("DELETE FROM tshirts")
        cursor.execute("DELETE FROM campers")
        print("Cleaned existing campers, checkins & tshirts tables in transaction.")

        # 3. First pass: group adults by Family Group to extract parent/guardian details
        family_guardians = {}
        for row in data_rows:
            if not row or len(row) < 3 or not row[2]:
                continue
            member_type = str(row[1]).strip() if len(row) > 1 and row[1] else 'Adult'
            name = str(row[2]).strip()
            phone = row[8] if len(row) > 8 else None
            email = row[9] if len(row) > 9 else None
            family_group = str(row[10]).strip() if len(row) > 10 and row[10] is not None else None
            
            if not family_group or not name:
                continue
                
            if family_group not in family_guardians:
                family_guardians[family_group] = { 'names': [], 'phone': None, 'email': None }
                
            if member_type == 'Adult':
                family_guardians[family_group]['names'].append(name)
                if phone and not family_guardians[family_group]['phone']:
                    family_guardians[family_group]['phone'] = str(phone).strip()
                if email and not family_guardians[family_group]['email']:
                    family_guardians[family_group]['email'] = str(email).strip()

        # 4. Second pass: insert campers
        inserted_count = 0
        tshirt_count = 0
        
        insert_sql = """
            INSERT INTO campers (
                first_name, last_name, age, gender, cabin_group, family_group,
                guardian_name, guardian_phone, allergies, registration_status, notes,
                kayaking, boat_tour
            ) VALUES (
                %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s
            )
        """
        
        for row in data_rows:
            if not row or len(row) < 3 or not row[2]:
                continue
            
            member_type = str(row[1]).strip() if len(row) > 1 and row[1] else 'Adult'
            name = str(row[2]).strip()
            raw_tshirt_size = str(row[3]).strip() if len(row) > 3 and row[3] else None
            child_age = row[4] if len(row) > 4 else None
            allergies = str(row[7]).strip() if len(row) > 7 and row[7] else None
            phone = row[8] if len(row) > 8 else None
            email = row[9] if len(row) > 9 else None
            family_group = str(row[10]).strip() if len(row) > 10 and row[10] is not None else None
            
            # Parse outdoor activities
            kayaking_val = 0
            try:
                if len(row) > 5 and row[5] is not None and str(row[5]).strip() != "":
                    kayaking_val = int(float(row[5]))
            except:
                pass

            boat_tour_val = 0
            try:
                if len(row) > 6 and row[6] is not None and str(row[6]).strip() != "":
                    boat_tour_val = int(float(row[6]))
            except:
                pass

            if not name:
                continue
                
            first_name, last_name = split_name(name)
            
            age = None
            if member_type == 'Child' and child_age is not None:
                try:
                    age = int(child_age)
                except ValueError:
                    pass
                    
            guardian_name = None
            guardian_phone = None
            
            if family_group and family_group in family_guardians:
                guardians = family_guardians[family_group]
                if guardians['names']:
                    guardian_name = " & ".join(guardians['names'])
                guardian_phone = guardians['phone']
                
            notes = f"Type: {member_type}"
            if member_type == 'Adult':
                guardian_name = "Self"
                guardian_phone = str(phone).strip() if phone else None
                
            cursor.execute(insert_sql, (
                first_name,
                last_name,
                age,
                None, # gender
                None, # cabin_group
                family_group,
                guardian_name,
                guardian_phone,
                allergies,
                'registered',
                notes,
                kayaking_val,
                boat_tour_val
            ))
            inserted_camper_id = cursor.lastrowid
            inserted_count += 1

            if raw_tshirt_size:
                cursor.execute(
                    "INSERT INTO tshirts (camper_id, attendee_name, tshirt_size, indian_size) VALUES (%s, %s, %s, %s)",
                    (inserted_camper_id, f"{first_name} {last_name}".strip(), raw_tshirt_size, "")
                )
                tshirt_count += 1
            
        conn.commit()
        print(f"Data migration finished! Successfully inserted {inserted_count} members and {tshirt_count} T-shirt sizes.")
    except Exception as e:
        conn.rollback()
        print(f"Error occurred during database operations: {e}")
        print("Database transaction has been rolled back. Existing database entries are untouched.")
    finally:
        cursor.close()
        conn.close()

if __name__ == "__main__":
    migrate()
