import os
import sys
import openpyxl
import pymysql
import urllib.parse as urlparse
from dotenv import load_dotenv

file_path = r"C:\Users\jenyj\OneDrive\MACBOOK\GCA\church-camp\GCA Tshirt Report.xlsx"
if not os.path.exists(file_path):
    file_path = "GCA Tshirt Report.xlsx"

if not os.path.exists(file_path):
    file_path = "GCA Camp Report.xlsx"

def normalize_size(val):
    if not val:
        return None
    val = str(val).strip()
    if not val:
        return None
    return val

def migrate():
    if not os.path.exists(file_path):
        print(f"Error: No T-Shirt Excel file found at: {file_path}")
        return

    print(f"Loading workbook: {file_path}...")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if "T-Shirts" in wb.sheetnames:
            sheet = wb["T-Shirts"]
        elif "Adults_Children" in wb.sheetnames:
            sheet = wb["Adults_Children"]
        else:
            sheet = wb.active

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError("T-Shirt Excel file has no data rows")
    except Exception as e:
        print(f"Error: Failed to load T-Shirt Excel workbook from {file_path}. Aborting.")
        print(f"Details: {e}")
        return

    # Connect to local MySQL database
    load_dotenv(os.path.join(os.path.dirname(__file__), '.env'))
    load_dotenv(os.path.join(os.path.dirname(__file__), 'backend', '.env'))
    
    db_url = os.environ.get("DATABASE_URL")
    db_user = os.environ.get("DB_USER", "")
    db_password = os.environ.get("DB_PASSWORD", "")
    db_host = os.environ.get("DB_HOST", "localhost")
    db_port = int(os.environ.get("DB_PORT", 3306))
    db_name = os.environ.get("DB_NAME", "churchcamp")

    if db_url:
        try:
            parsed_url = db_url.replace("mysql+pymysql://", "http://")
            url = urlparse.urlparse(parsed_url)
            db_user = url.username or db_user
            db_password = url.password or db_password
            db_host = url.hostname or db_host
            db_port = url.port or db_port
            db_name = url.path.lstrip('/') or db_name
        except Exception as e:
            print(f"Warning: Failed to parse DATABASE_URL from environment: {e}. Using default values.")

    conn = pymysql.connect(
        host=db_host, 
        user=db_user, 
        password=db_password, 
        port=db_port, 
        database=db_name
    )
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    matched_count = 0
    created_count = 0
    updated_count = 0
    unmatched_campers = []

    # Locate column indexes dynamically from the header row
    header = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    name_idx = -1
    tshirt_idx = -1
    indian_idx = -1

    for i, col in enumerate(header):
        if "name" in col and name_idx == -1:
            name_idx = i
        elif ("t-shirt" in col or "tshirt" in col or "size" in col) and "indian" not in col and tshirt_idx == -1:
            tshirt_idx = i
        elif "indian" in col and indian_idx == -1:
            indian_idx = i

    # Fallbacks if dynamic detection fails
    if name_idx == -1:
        name_idx = 1
    if tshirt_idx == -1:
        tshirt_idx = 3

    print(f"Using Columns -> Name: Col {name_idx+1}, US Size: Col {tshirt_idx+1}, Indian Size: Col {indian_idx+1 if indian_idx != -1 else 'N/A'}")

    # Process rows
    for idx, row in enumerate(rows[1:], start=1):
        if not row or len(row) <= name_idx or not row[name_idx]:
            continue
            
        name = str(row[name_idx]).strip()
        raw_us_size = row[tshirt_idx] if len(row) > tshirt_idx else None
        raw_indian_size = row[indian_idx] if indian_idx != -1 and len(row) > indian_idx else None
        
        if not name or name.lower() == "name":
            continue
            
        us_size = normalize_size(raw_us_size) or ""
        indian_size = normalize_size(raw_indian_size) or ""
            
        # Match camper in the local database by first_name + ' ' + last_name
        query = """
            SELECT id, first_name, last_name 
            FROM campers 
            WHERE LOWER(TRIM(CONCAT(first_name, ' ', last_name))) = LOWER(%s)
        """
        cursor.execute(query, (name,))
        db_campers = cursor.fetchall()
        
        matched_camper = None
        if len(db_campers) >= 1:
            matched_camper = db_campers[0]
        else:
            # First & last name fallback
            name_parts = name.split()
            if len(name_parts) >= 2:
                first = name_parts[0].lower()
                last = name_parts[-1].lower()
                cursor.execute("SELECT id, first_name, last_name FROM campers")
                all_campers = cursor.fetchall()
                for c in all_campers:
                    c_first = c["first_name"].strip().lower()
                    c_last = c["last_name"].strip().lower()
                    if first in c_first and last in c_last:
                        matched_camper = c
                        break

        if matched_camper:
            matched_count += 1
            camper_id = matched_camper["id"]
            full_name = f"{matched_camper['first_name']} {matched_camper['last_name']}"
            
            # Check if Tshirt record exists
            cursor.execute("SELECT id, tshirt_size, indian_size FROM tshirts WHERE camper_id = %s", (camper_id,))
            existing_tshirt = cursor.fetchone()
            
            if existing_tshirt:
                if existing_tshirt["tshirt_size"] != us_size or existing_tshirt["indian_size"] != indian_size:
                    cursor.execute(
                        "UPDATE tshirts SET tshirt_size = %s, indian_size = %s, attendee_name = %s WHERE id = %s",
                        (us_size, indian_size, full_name, existing_tshirt["id"])
                    )
                    updated_count += 1
            else:
                cursor.execute(
                    "INSERT INTO tshirts (camper_id, attendee_name, tshirt_size, indian_size) VALUES (%s, %s, %s, %s)",
                    (camper_id, full_name, us_size, indian_size)
                )
                created_count += 1
        else:
            unmatched_campers.append((name, raw_us_size, raw_indian_size))

    conn.commit()
    cursor.close()
    conn.close()

    print("\n" + "="*50)
    print("T-SHIRT LOCAL DB MIGRATION SUMMARY:")
    print(f"Total T-Shirt Excel rows processed: {len(rows) - 1}")
    print(f"Total matched campers in local DB: {matched_count}")
    print(f"Total T-shirts created: {created_count}")
    print(f"Total T-shirts updated: {updated_count}")
    print(f"Total unmatched campers: {len(unmatched_campers)}")

    if unmatched_campers:
        print("\nUnmatched campers from T-Shirt Excel:")
        for name, us_sz, ind_sz in unmatched_campers[:15]:
            print(f" - Name: {name}, US Size: {us_sz}, Indian Size: {ind_sz}")
    print("="*50)

if __name__ == "__main__":
    migrate()
